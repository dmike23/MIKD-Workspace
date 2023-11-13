import os

# hello

current_directory = os.getcwd()
print("Current working directory:", current_directory)

import openpyxl


def check_field_values(filename):
    results = []  # List to store the results

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(filename)

    # Select the Sheet1
    sheet = workbook['Sheet1']

    # Get the maximum number of rows in the sheet
    max_row = sheet.max_row

    for row in range(2, max_row + 1):
        case_id = sheet.cell(row=row, column=3).value

        # Check if CaseID exists in column A
        found = False
        for i in range(2, max_row + 1):
            if sheet.cell(row=i, column=1).value == case_id:
                field_value = sheet.cell(row=i, column=2).value
                if field_value is not None and "null" not in field_value and "remove section from JSON output" not in field_value:
                    results.append("Y")
                else:
                    results.append("N")
                found = True
                break

        if not found:
            results.append("N")

    return results


# Provide the filename of the Excel file
filename = "Book1.xlsx"

# Call the function to perform the checks
results = check_field_values(filename)

# Print the results
for result in results:
    print(result)
